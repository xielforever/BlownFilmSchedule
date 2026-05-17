import random
import datetime
import openpyxl
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from src.data_ingestion import BlownFilmDataIngestionPipeline
from src.config import INPUT_EXCEL_PATH

def generate_safe_orders():
    pipeline = BlownFilmDataIngestionPipeline()
    machines, orders, recipes_map, setup_mgr = pipeline.load_from_excel(INPUT_EXCEL_PATH)
    
    file_path = INPUT_EXCEL_PATH
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.worksheets[1]
    
    # 彻底清理之前追加的订单
    start_del_row = None
    for row in range(2, sheet.max_row + 1):
        val = sheet.cell(row=row, column=1).value
        if val == 'ORD-033':
            start_del_row = row
            break
            
    if start_del_row:
        sheet.delete_rows(start_del_row, sheet.max_row - start_del_row + 1)
        
    start_id = 33
    base_date = datetime.datetime(2026, 5, 17, 8, 0)
    
    product_types = list(recipes_map.keys())
    
    for i in range(200):
        o_id = f"ORD-{start_id + i:03d}"
        
        # 为了绝对安全，先随机挑一台机器，然后再生成符合它能力的订单
        m = random.choice(machines)
        
        # 挑选该机器能做的层数的产品
        valid_products = []
        for pt, mats in recipes_map.items():
            if len(mats) <= m.layer_structure:
                valid_products.append(pt)
                
        if not valid_products:
            valid_products = ['医药袋常规内衬膜'] # fallback
            
        ptype = random.choice(valid_products)
        
        # 宽度在机器能力范围内，偶尔给一些贴近极限的值
        if random.random() < 0.1:
            width = m.max_width  # 极宽挑战
        else:
            width = random.randint(m.min_width, m.max_width)
            
        # 厚度在机器能力范围内
        thickness = random.randint(m.min_thickness, m.max_thickness)
        
        # 极大订单或极小订单
        if random.random() < 0.1:
            qty = random.choice([500, m.hourly_output_kg * 48]) # 跑2天的大单
        else:
            qty = random.randint(2000, 15000)
            
        # 洁净室
        # 不能超出机器的洁净能力
        # Class_100K 比较脏，不能做 Class_10K 的单
        # 所以机器是 Class_100K，单子必须是 Class_100K 或 NO
        if m.cleanroom_level == 'Class_100K':
            cleanroom = random.choice(['Class_100K', 'NO'])
        else:
            cleanroom = random.choice(['Class_10K', 'Class_100K', 'NO'])
            
        if ptype == '临床试验加急样品膜':
            oclass = 'SAMPLE'
        else:
            oclass = random.choices(['NORMAL', 'URGENT'], weights=[0.8, 0.2])[0]
            
        cclass = 'VIP' if oclass == 'URGENT' or random.random() < 0.2 else 'STANDARD'
        
        order_date = base_date - datetime.timedelta(days=random.randint(1, 5), hours=random.randint(0, 23))
        
        if oclass == 'URGENT':
            due_date = base_date + datetime.timedelta(hours=random.randint(24, 72))
        else:
            due_date = base_date + datetime.timedelta(days=random.randint(3, 20), hours=random.randint(0, 23))
            
        corona = random.choice(['YES', 'NO'])
        core = random.choice([3, 6])
        
        mat_avail = 0
        if random.random() < 0.1:
            mat_avail = random.randint(1440, 2880)
            
        row_data = [
            o_id, ptype, width, thickness, qty, cleanroom, 
            order_date.strftime("%Y-%m-%d %H:%M"),
            due_date.strftime("%Y-%m-%d %H:%M"),
            cclass, oclass, corona, core, mat_avail
        ]
        
        sheet.append(row_data)
        
    wb.save(file_path)
    print(f"成功清理并重新生成追加 200 个绝不逾越物理边界的订单到 {file_path}")

if __name__ == '__main__':
    generate_safe_orders()
